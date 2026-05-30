"""Document reader tools for non-UTF-8 workspace formats."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from simplified_chatbot.tools.base import Tool, tool_parameters
from simplified_chatbot.tools.filesystem import _is_blocked_device, _resolve_path


_MAX_OUTPUT_CHARS = 128_000
_SUPPORTED_SUFFIXES = (".pdf", ".docx", ".xlsx")


def _validate_path(
    path: str | None,
    workspace: Path,
    allowed_dir: Path,
) -> tuple[Path, str | None]:
    """Resolve and validate a document path. Returns (resolved_path, error_message)."""
    if not path:
        return Path(), "Error: Unknown path"
    if _is_blocked_device(path):
        return Path(), f"Error: Reading {path} is blocked (device path that could hang)."
    fp = _resolve_path(path, workspace, allowed_dir)
    if _is_blocked_device(fp):
        return fp, f"Error: Reading {fp} is blocked (device path that could hang)."
    if not fp.exists():
        return fp, f"Error: File not found: {path}"
    if not fp.is_file():
        return fp, f"Error: Not a file: {path}"
    return fp, None


def _truncate_text(text: str, *, limit: int = _MAX_OUTPUT_CHARS, note: str) -> str:
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "\n\n" + note


def _parse_page_ranges(pages: str, page_count: int) -> list[int]:
    selected: set[int] = set()
    for part in pages.split(","):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            start_raw, end_raw = token.split("-", 1)
            start = int(start_raw)
            end = int(end_raw)
            if start < 1 or end < 1 or start > end:
                raise ValueError(f"Invalid page range: {token}")
            for page in range(start, end + 1):
                if page > page_count:
                    raise ValueError(
                        f"Requested page {page} is beyond end of document ({page_count} pages)",
                    )
                selected.add(page)
            continue
        page = int(token)
        if page < 1:
            raise ValueError(f"Invalid page number: {token}")
        if page > page_count:
            raise ValueError(
                f"Requested page {page} is beyond end of document ({page_count} pages)",
            )
        selected.add(page)

    if not selected:
        raise ValueError("pages must not be empty")
    return sorted(selected)


@tool_parameters(
    {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "The PDF path to read.",
            },
            "pages": {
                "type": "string",
                "description": "Optional page ranges such as '1-3,5'. Default reads all pages.",
            },
        },
        "required": ["path"],
    },
)
class ReadPdfTool(Tool):
    """Read text from PDF documents inside the workspace."""

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ) -> None:
        self._workspace = workspace.resolve() if workspace is not None else Path.cwd().resolve()
        self._allowed_dir = allowed_dir.resolve() if allowed_dir is not None else self._workspace

    @property
    def name(self) -> str:
        return "read_pdf"

    @property
    def description(self) -> str:
        return (
            "Read text from PDF files in the workspace. "
            "Use pages='1-3,5' to limit extraction when helpful."
        )

    async def execute(
        self,
        path: str | None = None,
        pages: str | None = None,
        **kwargs: Any,
    ) -> str:
        fp, error = _validate_path(path, self._workspace, self._allowed_dir)
        if error:
            return error
        return _extract_pdf(fp, path or str(fp), pages)


def _extract_pdf(fp: Path, display_path: str, pages: str | None) -> str:
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(fp))
        page_count = len(reader.pages)
        if page_count == 0:
            return f"(Empty PDF: {display_path})"
        selected_pages = (
            _parse_page_ranges(pages, page_count)
            if pages
            else list(range(1, page_count + 1))
        )

        chunks: list[str] = []
        for page_number in selected_pages:
            page_text = reader.pages[page_number - 1].extract_text() or ""
            normalized = page_text.replace("\r\n", "\n").strip()
            if not normalized:
                normalized = "[No extractable text on this page]"
            chunks.append(f"[Page {page_number}]\n{normalized}")

        result = "\n\n".join(chunks)
        result = _truncate_text(
            result,
            note=(
                f"(Output truncated after ~{_MAX_OUTPUT_CHARS} chars. "
                f"Refine pages=... to read a smaller PDF slice.)"
            ),
        )
        result += (
            f"\n\n(Extracted {len(selected_pages)} page(s) from {display_path}; "
            f"document has {page_count} page(s) total)"
        )
        return result
    except ImportError:
        return "Error: PDF support is not installed. Install 'pypdf' first."
    except ValueError as exc:
        return f"Error: {exc}"
    except PermissionError as exc:
        return f"Error: {exc}"
    except Exception as exc:
        return f"Error reading PDF: {exc}"


def _iter_docx_blocks(document: Any) -> list[str]:
    from docx.document import Document as DocumentType
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    if not isinstance(document, DocumentType):
        raise TypeError("Expected a python-docx Document")

    blocks: list[str] = []
    table_index = 0
    paragraph_index = 0
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            text = paragraph.text.strip()
            if not text:
                continue
            paragraph_index += 1
            blocks.append(f"[Paragraph {paragraph_index}]\n{text}")
            continue
        if isinstance(child, CT_Tbl):
            table_index += 1
            table = Table(child, document)
            rows: list[str] = []
            for row in table.rows:
                cells = [" ".join(cell.text.split()) for cell in row.cells]
                if any(cell for cell in cells):
                    rows.append(" | ".join(cells))
            if rows:
                blocks.append(f"[Table {table_index}]\n" + "\n".join(rows))
    return blocks


@tool_parameters(
    {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "The DOCX path to read.",
            },
        },
        "required": ["path"],
    },
)
class ReadDocxTool(Tool):
    """Read text and simple tables from DOCX files inside the workspace."""

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ) -> None:
        self._workspace = workspace.resolve() if workspace is not None else Path.cwd().resolve()
        self._allowed_dir = allowed_dir.resolve() if allowed_dir is not None else self._workspace

    @property
    def name(self) -> str:
        return "read_docx"

    @property
    def description(self) -> str:
        return (
            "Read text from DOCX files in the workspace, including paragraphs and simple tables."
        )

    async def execute(
        self,
        path: str | None = None,
        **kwargs: Any,
    ) -> str:
        fp, error = _validate_path(path, self._workspace, self._allowed_dir)
        if error:
            return error
        return _extract_docx(fp, path or str(fp))


def _extract_docx(fp: Path, display_path: str) -> str:
    try:
        from docx import Document

        document = Document(str(fp))
        blocks = _iter_docx_blocks(document)
        if not blocks:
            return f"(No extractable text in DOCX: {display_path})"

        result = "\n\n".join(blocks)
        result = _truncate_text(
            result,
            note=(
                f"(Output truncated after ~{_MAX_OUTPUT_CHARS} chars. "
                "Consider splitting the source document if you need a smaller slice.)"
            ),
        )
        result += f"\n\n(Extracted {len(blocks)} block(s) from {display_path})"
        return result
    except ImportError:
        return "Error: DOCX support is not installed. Install 'python-docx' first."
    except PermissionError as exc:
        return f"Error: {exc}"
    except Exception as exc:
        return f"Error reading DOCX: {exc}"


def _normalize_xlsx_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r\n", "\n").strip()
    return str(value)


@tool_parameters(
    {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "The XLSX path to read.",
            },
            "sheet": {
                "type": "string",
                "description": "Optional worksheet name. Default uses the active sheet.",
            },
            "range": {
                "type": "string",
                "description": "Optional A1 range such as 'A1:D20'. Default uses the used range.",
            },
        },
        "required": ["path"],
    },
)
class ReadXlsxTool(Tool):
    """Read worksheet data from XLSX files inside the workspace."""

    _DEFAULT_MAX_ROWS = 200

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ) -> None:
        self._workspace = workspace.resolve() if workspace is not None else Path.cwd().resolve()
        self._allowed_dir = allowed_dir.resolve() if allowed_dir is not None else self._workspace

    @property
    def name(self) -> str:
        return "read_xlsx"

    @property
    def description(self) -> str:
        return (
            "Read worksheet values from XLSX files in the workspace. "
            "Optionally target one sheet and A1 range."
        )

    async def execute(
        self,
        path: str | None = None,
        sheet: str | None = None,
        range: str | None = None,
        **kwargs: Any,
    ) -> str:
        fp, error = _validate_path(path, self._workspace, self._allowed_dir)
        if error:
            return error
        return _extract_xlsx(fp, path or str(fp), sheet, range, self._DEFAULT_MAX_ROWS)


_XLSX_DEFAULT_MAX_ROWS = 200


def _extract_xlsx(
    fp: Path,
    display_path: str,
    sheet: str | None,
    range: str | None,
    max_rows: int = _XLSX_DEFAULT_MAX_ROWS,
) -> str:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        workbook = load_workbook(filename=str(fp), read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet else workbook.active

        if range:
            min_col, min_row, max_col, max_row = range_boundaries(range)
            range_label = range
        else:
            min_col, min_row = 1, 1
            max_col = max(1, worksheet.max_column or 1)
            max_row = max(1, worksheet.max_row or 1)
            range_label = worksheet.calculate_dimension()

        lines: list[str] = [f"[Sheet: {worksheet.title}]", f"[Range: {range_label}]"]
        row_count = 0
        truncated = False
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ),
            start=min_row,
        ):
            row_count += 1
            if row_count > max_rows:
                truncated = True
                break
            values = [_normalize_xlsx_cell(value) for value in row]
            lines.append(f"{row_index}| " + "\t".join(values))

        result = "\n".join(lines)
        result = _truncate_text(
            result,
            note=(
                f"(Output truncated after ~{_MAX_OUTPUT_CHARS} chars. "
                "Narrow the sheet or range to read a smaller slice.)"
            ),
        )
        if truncated:
            result += (
                f"\n\n(Stopped after {max_rows} rows. "
                "Provide a smaller range=... for a focused read.)"
            )
        result += f"\n\n(Extracted {min(row_count, max_rows)} row(s) from {display_path})"
        return result
    except ImportError:
        return "Error: XLSX support is not installed. Install 'openpyxl' first."
    except KeyError:
        return f"Error: Worksheet not found: {sheet}"
    except ValueError as exc:
        return f"Error: {exc}"
    except PermissionError as exc:
        return f"Error: {exc}"
    except Exception as exc:
        return f"Error reading XLSX: {exc}"


@tool_parameters(
    {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "Path to a .pdf, .docx, or .xlsx file. Format detected by extension.",
            },
            "pages": {
                "type": "string",
                "description": (
                    "PDF only: page ranges such as '1-3,5'. Default reads all pages. "
                    "Ignored for non-PDF formats."
                ),
            },
            "sheet": {
                "type": "string",
                "description": (
                    "XLSX only: worksheet name. Default uses the active sheet. "
                    "Ignored for non-XLSX formats."
                ),
            },
            "range": {
                "type": "string",
                "description": (
                    "XLSX only: A1 range such as 'A1:D20'. Default uses the used range. "
                    "Ignored for non-XLSX formats."
                ),
            },
        },
        "required": ["path"],
    },
)
class ReadDocumentTool(Tool):
    """Read text from PDF, DOCX, or XLSX files; dispatches by file extension."""

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ) -> None:
        self._workspace = workspace.resolve() if workspace is not None else Path.cwd().resolve()
        self._allowed_dir = allowed_dir.resolve() if allowed_dir is not None else self._workspace

    @property
    def name(self) -> str:
        return "read_document"

    @property
    def description(self) -> str:
        return (
            "Read text from office documents (.pdf, .docx, .xlsx) in the workspace. "
            "Dispatches by file extension. "
            "Use 'pages' for PDF page selection; 'sheet'/'range' for XLSX targeting. "
            "Use read_file for UTF-8 text files."
        )

    async def execute(
        self,
        path: str | None = None,
        pages: str | None = None,
        sheet: str | None = None,
        range: str | None = None,
        **kwargs: Any,
    ) -> str:
        fp, error = _validate_path(path, self._workspace, self._allowed_dir)
        if error:
            return error
        display_path = path or str(fp)
        suffix = fp.suffix.lower()
        if suffix == ".pdf":
            return _extract_pdf(fp, display_path, pages)
        if suffix == ".docx":
            return _extract_docx(fp, display_path)
        if suffix == ".xlsx":
            return _extract_xlsx(fp, display_path, sheet, range)
        return (
            f"Error: Unsupported document format '{suffix or '(none)'}'. "
            f"Supported: {', '.join(_SUPPORTED_SUFFIXES)}. "
            "Use read_file for UTF-8 text."
        )
